import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_CSV = BASE_DIR / "device_data" / "success.csv"
OUTPUT_XLSX = BASE_DIR / "device_data" / "scheduler_report.xlsx"

TARGET_MODES = [
    "ForceCharge",
    "ForceDischarge",
    "ForceCharge(BAT)",
    "ForceDischarge(BAT)",
]


def format_time(hour, minute):
    hour = int(hour or 0)
    minute = int(minute or 0)
    return f"{hour:02d}:{minute:02d}"


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def write_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def load_rows():
    detail_rows = []
    summary_rows = {}
    stats = Counter()
    total_rows = 0
    processed_rows = 0
    with_groups = 0
    skipped_blank_json = 0
    skipped_header_rows = 0
    skipped_invalid_json = 0

    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for line_number, row in enumerate(reader, start=2):
            total_rows += 1
            response_json = (row.get("response_json") or "").strip()
            if not response_json:
                skipped_blank_json += 1
                continue
            if response_json == "response_json":
                skipped_header_rows += 1
                continue

            try:
                payload = json.loads(response_json)
            except json.JSONDecodeError:
                skipped_invalid_json += 1
                print(
                    f"跳过无效 JSON: line={line_number}, device_sn={row.get('device_sn', '')}"
                )
                continue

            processed_rows += 1
            result = payload.get("result", {})
            groups = result.get("groups", []) or []

            if groups:
                with_groups += 1

            summary_entry = summary_rows.setdefault(
                row["device_sn"],
                {
                    "device_id": row["device_id"],
                    "device_sn": row["device_sn"],
                    "protocol_version": row["protocol_version"],
                    "master_version": row["master_version"],
                    "source_column_5": row.get("source_column_5", ""),
                    "source_column_6": row.get("source_column_6", ""),
                    "force_charge_periods": [],
                    "force_discharge_periods": [],
                    "force_charge_bat_periods": [],
                    "force_discharge_bat_periods": [],
                },
            )

            for group_index, group in enumerate(groups, start=1):
                work_mode = group.get("workMode") or group.get("workmode") or ""
                stats[work_mode] += 1
                if work_mode not in TARGET_MODES:
                    continue

                extra = group.get("extraParam", {}) or {}
                start_time = format_time(group.get("startHour"), group.get("startMinute"))
                end_time = format_time(group.get("endHour"), group.get("endMinute"))
                period = f"{start_time}-{end_time}"

                detail_rows.append(
                    {
                        "device_id": row["device_id"],
                        "device_sn": row["device_sn"],
                        "protocol_version": row["protocol_version"],
                        "master_version": row["master_version"],
                        "source_column_5": row.get("source_column_5", ""),
                        "source_column_6": row.get("source_column_6", ""),
                        "work_mode": work_mode,
                        "start_time": start_time,
                        "end_time": end_time,
                        "period": period,
                        "fd_pwr": extra.get("fdPwr"),
                        "fd_soc": extra.get("fdSoc"),
                        "min_soc_on_grid": extra.get("minSocOnGrid"),
                        "max_soc": extra.get("maxSoc"),
                        "pv_limit": extra.get("pvLimit"),
                        "import_limit": extra.get("importLimit"),
                        "export_limit": extra.get("exportLimit"),
                        "reactive_power": extra.get("reactivePower"),
                        "group_index": group_index,
                    }
                )

                if work_mode == "ForceCharge":
                    summary_entry["force_charge_periods"].append(period)
                elif work_mode == "ForceDischarge":
                    summary_entry["force_discharge_periods"].append(period)
                elif work_mode == "ForceCharge(BAT)":
                    summary_entry["force_charge_bat_periods"].append(period)
                elif work_mode == "ForceDischarge(BAT)":
                    summary_entry["force_discharge_bat_periods"].append(period)
    skipped_rows = skipped_blank_json + skipped_header_rows + skipped_invalid_json
    return (
        detail_rows,
        summary_rows,
        stats,
        total_rows,
        processed_rows,
        with_groups,
        skipped_rows,
        skipped_blank_json,
        skipped_header_rows,
        skipped_invalid_json,
    )


def build_workbook(
    detail_rows,
    summary_rows,
    stats,
    total_rows,
    processed_rows,
    with_groups,
    skipped_rows,
    skipped_blank_json,
    skipped_header_rows,
    skipped_invalid_json,
):
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "明细"
    detail_headers = [
        "device_id",
        "device_sn",
        "protocol_version",
        "master_version",
        "source_column_5",
        "source_column_6",
        "work_mode",
        "start_time",
        "end_time",
        "period",
        "fd_pwr",
        "fd_soc",
        "min_soc_on_grid",
        "max_soc",
        "pv_limit",
        "import_limit",
        "export_limit",
        "reactive_power",
        "group_index",
    ]
    write_header(ws_detail, detail_headers)
    for row in detail_rows:
        ws_detail.append([row.get(header) for header in detail_headers])
    auto_fit_columns(ws_detail)

    ws_summary = wb.create_sheet("汇总")
    summary_headers = [
        "device_id",
        "device_sn",
        "protocol_version",
        "master_version",
        "source_column_5",
        "source_column_6",
        "force_charge_periods",
        "force_discharge_periods",
        "force_charge_bat_periods",
        "force_discharge_bat_periods",
    ]
    write_header(ws_summary, summary_headers)
    for item in summary_rows.values():
        if not any(
            [
                item["force_charge_periods"],
                item["force_discharge_periods"],
                item["force_charge_bat_periods"],
                item["force_discharge_bat_periods"],
            ]
        ):
            continue
        ws_summary.append(
            [
                item["device_id"],
                item["device_sn"],
                item["protocol_version"],
                item["master_version"],
                item["source_column_5"],
                item["source_column_6"],
                ", ".join(item["force_charge_periods"]),
                ", ".join(item["force_discharge_periods"]),
                ", ".join(item["force_charge_bat_periods"]),
                ", ".join(item["force_discharge_bat_periods"]),
            ]
        )
    auto_fit_columns(ws_summary)

    ws_stats = wb.create_sheet("统计")
    write_header(ws_stats, ["metric", "value"])
    ws_stats.append(["csv_rows", total_rows])
    ws_stats.append(["processed_rows", processed_rows])
    ws_stats.append(["rows_with_groups", with_groups])
    ws_stats.append(["detail_rows", len(detail_rows)])
    ws_stats.append(["skipped_rows", skipped_rows])
    ws_stats.append(["skipped_blank_json", skipped_blank_json])
    ws_stats.append(["skipped_header_rows", skipped_header_rows])
    ws_stats.append(["skipped_invalid_json", skipped_invalid_json])
    for mode in TARGET_MODES:
        ws_stats.append([mode, stats.get(mode, 0)])
    auto_fit_columns(ws_stats)

    return wb


def main():
    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"文件不存在: {INPUT_CSV}")

    (
        detail_rows,
        summary_rows,
        stats,
        total_rows,
        processed_rows,
        with_groups,
        skipped_rows,
        skipped_blank_json,
        skipped_header_rows,
        skipped_invalid_json,
    ) = load_rows()
    workbook = build_workbook(
        detail_rows,
        summary_rows,
        stats,
        total_rows,
        processed_rows,
        with_groups,
        skipped_rows,
        skipped_blank_json,
        skipped_header_rows,
        skipped_invalid_json,
    )
    workbook.save(OUTPUT_XLSX)

    print(f"输入文件: {INPUT_CSV}")
    print(f"输出文件: {OUTPUT_XLSX}")
    print(f"CSV 行数: {total_rows}")
    print(f"有效 JSON 行数: {processed_rows}")
    print(f"有调度分组的设备数: {with_groups}")
    print(f"导出明细行数: {len(detail_rows)}")
    print(f"跳过行数: {skipped_rows}")
    print(f"跳过空 JSON 行数: {skipped_blank_json}")
    print(f"跳过重复表头行数: {skipped_header_rows}")
    print(f"跳过无效 JSON 行数: {skipped_invalid_json}")
    for mode in TARGET_MODES:
        print(f"{mode}: {stats.get(mode, 0)}")


if __name__ == "__main__":
    main()
